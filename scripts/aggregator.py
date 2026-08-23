import pandas as pd
import requests
from datetime import datetime, timedelta
import os
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import PatternFill

BASE_URL = "https://api.polygon.io/v2/aggs/ticker"

# Function to Download Data from Polygon
def fetch_polygon_data(ticker, start_date, end_date, API_KEY):
    url = f"{BASE_URL}/{ticker}/range/1/day/{start_date}/{end_date}?apiKey={API_KEY}"
    response = requests.get(url)
    if response.status_code != 200:
        print(f"Failed to fetch data for {ticker}: {response.status_code}")
        return pd.DataFrame()

    data = response.json()
    if "results" not in data or not data["results"]:
        print(f"No data for {ticker}")
        return pd.DataFrame()

    df = pd.DataFrame(data["results"])
    df['Date'] = pd.to_datetime(df['t'], unit='ms').dt.date
    df.rename(columns={"c": "Close", "l": "Low", "h": "High"}, inplace=True)
    df = df[['Date', 'Close', 'Low', 'High']]
    return df

 
def calculate_metrics(df):
    df['Diff_Daily'] = df['Close'].pct_change(1) * 100
    df['Diff_Weekly'] = df['Close'].pct_change(5) * 100
    df['Diff_Monthly'] = df['Close'].pct_change(21) * 100
    
    # 52 Week Low and High - use entire available data for max/min
    df['52wl'] = (df['Close'].min() - df['Close']) / df['Close'].min() * 100 
    df['52wh'] = (df['Close'].max() - df['Close']) / df['Close'].max() * 100   
    return df.tail(1)   

# Create Data Directory
data_dir = "sector_data"
if not os.path.exists(data_dir):
    os.makedirs(data_dir)

def aggregate_and_rank_groups(GROUPS, api_key, start_date_str, end_date_str, output_file):
    """
    Aggregates and ranks groups, returning a dictionary of DataFrames.

    Args:
        GROUPS (dict): Dictionary of groups and their sectors/tickers.
        api_key (str): Polygon API key.
        start_date_str (str): Start date in YYYY-MM-DD format.
        end_date_str (str): End date in YYYY-MM-DD format.
        output_file (str): Path to save the Excel file.

    Returns:
        dict: Dictionary of DataFrames for each group.
    """
    print("Step 2-------Aggregating and Ranking---------")
    results = {group: [] for group in GROUPS}
    dataframes = {}

    for group, sectors in GROUPS.items():
        for sector, tickers in sectors.items():
            tickers = [tickers] if isinstance(tickers, str) else tickers
            for ticker in tickers:
                df = fetch_polygon_data(ticker, start_date_str, end_date_str, api_key)
                if df.empty:
                    continue
                df_metrics = calculate_metrics(df)
                results[group].append({
                    "Rank": None,
                    "Sector": sector,
                    "Ticker": ticker,
                    "Diff_Daily": round(df_metrics['Diff_Daily'].values[0], 1),
                    "Diff_Weekly": round(df_metrics['Diff_Weekly'].values[0], 1),
                    "Diff_Monthly": round(df_metrics['Diff_Monthly'].values[0], 1),
                    "52wl": round(df_metrics['52wl'].values[0], 1),
                    "52wh": round(df_metrics['52wh'].values[0], 1)
                })

        # Create DataFrame for the group
        df_group = pd.DataFrame(results[group])
        if not df_group.empty:
            df_group = df_group.sort_values(by="Diff_Daily", ascending=False).reset_index(drop=True)
            df_group['Rank'] = df_group.index + 1
            dataframes[group] = df_group

    # Save to Excel
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        workbook = writer.book
        bold_font = Font(bold=True)
        start_row = 0

        for group, data in results.items():
            df_group = pd.DataFrame(data)
            if df_group.empty:
                continue
            df_group = df_group.sort_values(by="Diff_Daily", ascending=False).reset_index(drop=True)
            df_group['Rank'] = df_group.index + 1

            # Add group header
            df_header = pd.DataFrame([[None, f"{group}", None]], columns=["Rank", "Sector", "Ticker"])
            df_header.to_excel(writer, sheet_name="Ranked_Data", index=False, startrow=start_row, header=False)
            df_group.to_excel(writer, sheet_name="Ranked_Data", index=False, startrow=start_row + 1)

            # Style group header
            start_row += len(df_group) + 3
            sheet = writer.sheets["Ranked_Data"]
            cell = sheet.cell(row=start_row - len(df_group) - 2, column=2)
            cell.font = bold_font

           
            rank_start = start_row - len(df_group)
            rank_end = start_row - 1
            rank_range = f"A{rank_start}:A{rank_end}"
            color_scale = ColorScaleRule(
                start_type="min", start_color="FF63BE7B",
                mid_type="percentile", mid_color="FFFFFF99", mid_value=50,
                end_type="max", end_color="FFFF6666"
            )
            sheet.conditional_formatting.add(rank_range, color_scale)

        
            low_start_col = "G"
            high_start_col = "H"
            low_range = f"{low_start_col}{rank_start}:{low_start_col}{rank_end}"
            high_range = f"{high_start_col}{rank_start}:{high_start_col}{rank_end}"
            data_bar_low = DataBarRule(
                start_type="num", start_value=0,
                end_type="max", color="63BE7B", showValue=True
            )
            data_bar_high = DataBarRule(
                start_type="num", start_value=0,
                end_type="max", color="FF6666", showValue=True
            )
            sheet.conditional_formatting.add(low_range, data_bar_low)
            sheet.conditional_formatting.add(high_range, data_bar_high)

            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    if not cell.has_style:
                        cell.fill = white_fill

    print(f"Excel file saved as {output_file}")
    return dataframes
